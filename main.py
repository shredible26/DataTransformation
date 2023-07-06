import pandas as pd
import openpyxl

# Step 1 (isolates the top table)

excel_file = "/Users/shreyvarma/Downloads/DynPro/Intern_Data_sheet_1.xlsx"
df = pd.read_excel(excel_file)

extracted_data = df.iloc[0:10, 6:18].copy()

print("Column labels in the extracted data:")
print(extracted_data.columns)
columns_to_remove = [1, 2, 4, 5, 6, 8, 9, 10]
extracted_data.drop(extracted_data.columns[columns_to_remove], axis=1, inplace=True)
output_file = "/Users/shreyvarma/Downloads/DynPro/Top_Table_Transition_1.xlsx"
extracted_data.to_excel(output_file, index=False)

# Step 2 (Converts top table to horizontal headers)

excel_file = "/Users/shreyvarma/Downloads/DynPro/Top_Table_Transition_1.xlsx"
df = pd.read_excel(excel_file)

num_rows = df.shape[0]
num_cols = df.shape[1]

modified_data = pd.DataFrame()

for col in range(1, num_cols):
    col_num = col

    for row in range(num_rows):
        row_name = df.iloc[row, 0]

        new_col_title = f"Interval_{col_num}_{row_name}"
        cell_value = df.iloc[row, col]
        modified_data[new_col_title] = [cell_value]

output_file = "/Users/shreyvarma/Downloads/DynPro/Top_Table_Transition_2.xlsx"
modified_data.to_excel(output_file, index=False)

print("Done!")

# Step 3 (initializes left table, horizontally)

# Read the Excel file
excel_file = "/Users/shreyvarma/Downloads/DynPro/Intern_Data_Left_Table.xlsx"
df = pd.read_excel(excel_file)

# Create a new DataFrame for the modified table
modified_data = pd.DataFrame()

# Process each row in the original table
for row in range(df.shape[0]):
    # Initialize an empty list to store non-empty cell values in the row
    cell_values = []

    # Traverse horizontally across the row
    for col in range(df.shape[1]):
        cell_value = df.iloc[row, col]
        if pd.notnull(cell_value):
            cell_values.append(str(cell_value))

    # Create the column title by joining the non-empty cell values with an underscore
    column_title = '_'.join(cell_values)

    # Add the column title as a new column to the modified table
    modified_data[column_title] = ""

# Shift the data one column to the right
modified_data = modified_data.shift(1, axis=1)

# Save the modified table to a new Excel file
output_file = "/Users/shreyvarma/Downloads/DynPro/Left_Table_Transition_1.xlsx"
modified_data.to_excel(output_file, index=False)

print("Done!")

# WORKING, modifies Left_Table_Transition_1

# Step 4 (Left table horizontal to verticle)

# Read the Excel file
excel_file = "/Users/shreyvarma/Downloads/DynPro/Left_Table_Transition_1.xlsx"
df = pd.read_excel(excel_file)

# Transpose the DataFrame to swap rows and columns
transposed_df = df.T

# Reset the index to create a new row index starting from 1
transposed_df = transposed_df.reset_index()

# Rename the columns to set the first column as 'A1', second column as 'A2', and so on
transposed_df.columns = ['A' + str(i+1) for i in range(transposed_df.shape[1])]

# Save the transposed table to a new Excel file
output_file = "/Users/shreyvarma/Downloads/DynPro/Left_Table_Transition_2.xlsx"
transposed_df.to_excel(output_file, index=False)

print("done!")

# Step 5 (Copying center table + cleaning it)

# Read the Excel file
input_file = "/Users/shreyvarma/Downloads/DynPro/Intern_Data_sheet_1.xlsx"
df = pd.read_excel(input_file)

# Extract the desired data from columns I to T and rows 17 to 54
extracted_data = df.iloc[16:53, 8:20].copy()

# Remove columns E to I from the extracted data
columns_to_remove = extracted_data.columns[4:9]
extracted_data.drop(columns_to_remove, axis=1, inplace=True)

# Remove the first three rows from the extracted data
extracted_data = extracted_data.iloc[2:]

# Reset the index of the extracted data
extracted_data.reset_index(drop=True, inplace=True)

# Save the extracted data to a new Excel file
output_file = "/Users/shreyvarma/Downloads/DynPro/Center_Table_Transition_1.xlsx"
extracted_data.to_excel(output_file, index=False)

print("Done!")

# Step 6 (Combining the left and middle tables)

# Read the Excel files
table1_file = "/Users/shreyvarma/Downloads/DynPro/Final/Left_Table_Transition_FINAL.xlsx"
table2_file = "/Users/shreyvarma/Downloads/DynPro/Final/Center_Table_Transition_FINAL.xlsx"

# Read table 1 (assuming it has only one column A)
table1 = pd.read_excel(table1_file, usecols="A")

# Read table 2 (assuming it has columns A to G)
table2 = pd.read_excel(table2_file, usecols="A:G")

# Join table 2 to table 1 starting from Column B
combined_table = pd.concat([table1, table2], axis=1)

# Save the combined table to a new Excel file
output_file = "/Users/shreyvarma/Downloads/DynPro/Left_And_Middle.xlsx"
combined_table.to_excel(output_file, index=False)

print("Tables joined and saved successfully!")

# Step 7 (Deleting empty rows)

# Read the Excel file
excel_file = "/Users/shreyvarma/Downloads/DynPro/Final/Left_And_Middle.xlsx"
df = pd.read_excel(excel_file)

# Define the rows to be deleted
rows_to_delete = [5, 17, 12, 16, 19, 24, 30, 34]

# Delete the specified rows
df.drop(rows_to_delete, inplace=True)

# Save the modified table to a new Excel file
output_file = "/path/to/new_table.xlsx"
df.to_excel(output_file, index=False)

print("Done!")

# Step 8 (Combining all tables to create final table!)

# Read the first Excel file (Table 1)
table1_file = "/Users/shreyvarma/Downloads/DynPro/Final/Top_Table_Transition_FINAL.xlsx"
df_table1 = pd.read_excel(table1_file)

# Read the second Excel file (Table 2)
table2_file = "/Users/shreyvarma/Downloads/DynPro/Final/Left_And_Middle_FINAL.xlsx"
df_table2 = pd.read_excel(table2_file)

# Attach Table 2 to the right of Table 1
df_combined = pd.concat([df_table1, df_table2], axis=1)

# Save the combined table to a new Excel file
output_file = "/Users/shreyvarma/Downloads/DynPro/Final/Final_Table.xlsx"
df_combined.to_excel(output_file, index=False)

print("Done!")

import pandas as pd

# Read the Excel file
excel_file = "/Users/shreyvarma/Downloads/DynPro/Final/Final_Table_FINAL.xlsx"
df = pd.read_excel(excel_file)

# Check the data types of the columns
data_types = df.dtypes

# Print the data types of each column
for column, dtype in data_types.iteritems():
    print(f"Column {column} has data type: {dtype}")

# Determine if values are strings or integers
for column, dtype in data_types.iteritems():
    if dtype == object:
        print(f"Column {column} contains strings")
    elif dtype == int:
        print(f"Column {column} contains integers")
    else:
        print("neither")

# Step 9 (Manually changing data types)

# Load the Excel file
workbook = openpyxl.load_workbook('/Users/shreyvarma/Downloads/DynPro/Final/Final_Table_FINAL.xlsx')

# Select the active sheet (or specify the sheet name)
sheet = workbook.active

# Iterate through columns A to R (inclusive)
for column in sheet.iter_cols(min_col=1, max_col=20):
    for cell in column:
        cell.number_format = '@'  # Set number format to text

# Iterate through columns S to Z (inclusive)
for column in sheet.iter_cols(min_col=21, max_col=26):
    for cell in column:
        cell.number_format = '0.00'  # Set number format to two decimal places

# Save the modified workbook
workbook.save('/Users/shreyvarma/Downloads/DynPro/Final/Final_Table_With_Data.xlsx')